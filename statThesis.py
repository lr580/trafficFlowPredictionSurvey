# 统计具体论文笔记.md里的基本信息汇总成表格
# 信息获取
with open('具体论文笔记.md', encoding='utf8') as f:
    text = f.read()
import re
pattern = r'### *(T-\d+) *([^\n]*)\s*CCF-(\w+)\s*CI-(\d+)\s*DT-(\d+) *([^\n]*)'
results = re.findall(pattern, text, re.S | re.M)

titles = ['编号', '模型名', 'CCF等级', '引用数', '年份', '问题类型']
thesis = []
distinctTarget = set()
for tid, abbr, ccf, ci, dt, target in results:
    target = sorted([i[4:] for i in target.split()])
    distinctTarget |= set(target)
    targets = '、'.join(target)
    ci, dt = int(ci), int(dt)
    ccf = ccf if ccf.strip() != 'NONE' else '无'
    thesis.append([tid, abbr, ccf, ci, dt, targets])
# checks:
# print(results)
# print(set([len(i) for i in results]))
# print(thesis['T-109'])
# print(thesis['T-44'])
# print(distinctTarget)

# 生成统计报告
def cmp(u):
    return ord(u[2]), -u[3] # 按CCF等级和引用数排序
thesis.sort(key = cmp)
report = ''
import pandas as pd
import numpy as np
df = pd.DataFrame(thesis, columns=titles)
report += f'共有论文{df.shape[0]}篇\n\n'
for ccf in df['CCF等级'].unique():
    report += f'CCF-{ccf}类论文共{len(df[df["CCF等级"] == ccf])}篇\n'
report += '\n'
for dt in np.sort(df['年份'].unique())[::-1]:
    report += f'{dt}年论文共{len(df[df["年份"] == dt])}篇\n'
report += '\n'
for target in sorted(distinctTarget):
    report += f'{target}类问题共{len(df[df["问题类型"].str.contains(target)])}篇\n'
with open('论文统计报告.txt', 'w', encoding='UTF8') as f:
    f.write(report)

# 表格填充
from openpyxl import *
wb = Workbook()
sh = wb['Sheet']
sh.title = '论文统计表'
for i, title in enumerate(titles):
    sh.cell(row = 1, column = i+1).value = title
for r, row in enumerate(thesis):
    for c, col in enumerate(row):
        sh.cell(row = r+2, column = c+1).value = col
        
# 样式调整
from openpyxl.styles import Alignment
# 自动调整列宽
for col in sh.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列字母
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2  # 添加一些空白边距
    sh.column_dimensions[column].width = adjusted_width
# 设置居中对齐
for row in sh.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
# 自动筛选
sh.auto_filter.ref = "A1:F1"

wb.save('论文统计表.xlsx')